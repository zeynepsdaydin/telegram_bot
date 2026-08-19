import json
import os

from database import log_api_call
from openpyxl import load_workbook

EXCEL_FILE = "MUKAYESE RAPORU GENEL.xlsx"
cached_data = None
_last_mtime = 0


def tr_lower(text: str) -> str:
    if not text:
        return ""
    trans = str.maketrans(
        {"İ": "i", "I": "ı", "Ğ": "ğ", "Ü": "ü", "Ş": "ş", "Ö": "ö", "Ç": "ç"}
    )
    return str(text).translate(trans).lower()


def normalize_text(text: str) -> str:
    text = tr_lower(text)
    trans = str.maketrans(
        {"ı": "i", "g": "g", "u": "u", "s": "s", "o": "o", "c": "c"}
    )
    return text.translate(trans)


def load_excel_to_memory():
    global cached_data, _last_mtime

    base_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(base_dir, EXCEL_FILE)

    if not os.path.exists(excel_path):
        return None

    try:
        current_mtime = os.path.getmtime(excel_path)
        # Dosya değişmediyse ve hafızada varsa tekrar diskten okuma
        if cached_data is not None and current_mtime == _last_mtime:
            return cached_data

        wb = load_workbook(excel_path, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        all_rows = list(sheet.iter_rows(values_only=True))

        prod_idx, unit_idx, qty_idx, link_idx, reason_idx = 2, 3, 4, 5, 6
        header_row_idx = 0
        row_3 = []

        for r_idx, row in enumerate(all_rows[:10]):
            row_str = [str(cell).upper().replace("İ", "I") for cell in row if cell]
            row_concat = " ".join(row_str)
            if "URUN ADI" in row_concat or "SIRA NO" in row_concat:
                header_row_idx = r_idx
                if r_idx > 0:
                    row_3 = all_rows[r_idx - 1]
                for c_idx, cell in enumerate(row):
                    if cell:
                        c_upper = str(cell).upper().replace("İ", "I").strip()
                        if "URUN ADI" in c_upper:
                            prod_idx = c_idx
                        elif "BIRIM" in c_upper and "BIRIM FIYAT" not in c_upper:
                            unit_idx = c_idx
                        elif "MIKTAR" in c_upper:
                            qty_idx = c_idx
                        elif "LINK" in c_upper:
                            link_idx = c_idx
                        elif "SEBEP" in c_upper:
                            reason_idx = c_idx
                break

        price_columns = {}
        if header_row_idx < len(all_rows):
            header_row = all_rows[header_row_idx]
            last_found_company = "Belirtilmeyen Firma"

            for idx, cell in enumerate(header_row):
                if (
                    idx < len(row_3)
                    and row_3[idx] is not None
                    and str(row_3[idx]).strip()
                ):
                    last_found_company = str(row_3[idx]).strip()

                if cell:
                    c_upper = str(cell).upper().replace("İ", "I").strip()
                    if "BIRIM FIYAT" in c_upper or "FIYAT" in c_upper:
                        price_columns[idx] = last_found_company

        temp_data = []
        for row in all_rows[header_row_idx + 1 :]:
            if not row or len(row) <= prod_idx or row[prod_idx] is None:
                continue

            urun_adi_val = str(row[prod_idx]).strip()
            if (
                not urun_adi_val
                or "URUN ADI" in urun_adi_val.upper()
                or urun_adi_val.isdigit()
            ):
                continue

            fiyat_teklifleri = []
            for col_idx, f_name in price_columns.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    val = str(row[col_idx]).strip()
                    if (
                        val
                        and val.lower() != "none"
                        and val not in ["0", "0.0", "stok yok"]
                    ):
                        if not val.endswith("TL") and "₺" not in val:
                            val = f"{val} TL"
                        fiyat_teklifleri.append(f"{f_name}: {val}")

            fiyat_ozeti = (
                ", ".join(fiyat_teklifleri)
                if fiyat_teklifleri
                else "Fiyat belirtilmedi / Stok Yok"
            )

            temp_data.append({
                "urun_adi": urun_adi_val,
                "birim": (
                    str(row[unit_idx]).strip()
                    if len(row) > unit_idx and row[unit_idx] is not None
                    else "Belirtilmedi"
                ),
                "miktar": (
                    str(row[qty_idx]).strip()
                    if len(row) > qty_idx and row[qty_idx] is not None
                    else "Belirtilmedi"
                ),
                "link": (
                    str(row[link_idx]).strip()
                    if len(row) > link_idx and row[link_idx] is not None
                    else "Belirtilmedi"
                ),
                "sebep": (
                    str(row[reason_idx]).strip()
                    if len(row) > reason_idx and row[reason_idx] is not None
                    else "Belirtilmedi"
                ),
                "fiyat_listesi": fiyat_ozeti,
            })

        cached_data = temp_data
        _last_mtime = current_mtime
        return cached_data
    except (OSError, ValueError, KeyError) as e:
        print(f"Excel yükleme hatası: {e!s}")
        return None


load_excel_to_memory()


def format_material_response(materials: list, query: str = "") -> str:
    """Malzeme listesini şık bir Markdown tablosuna dönüştürür."""
    if not materials:
        return f"Aradığınız malzeme ({query}) mukayese raporunda bulunamadı."

    header = "### 📦 Mukayese Raporu Malzeme Listesi\n\n"
    table_head = "| Ürün Adı | Miktar / Birim | Fiyat Bilgisi | Kullanım Amacı / Sebep | Satıcı / Link |\n"
    table_divider = "| :--- | :---: | :---: | :--- | :---: |\n"

    rows = []
    for item in materials:
        name = item.get("urun_adi", "Bilinmiyor")
        qty = f"{item.get('miktar', '-')} {item.get('birim', '')}".strip()
        price = item.get("fiyat_listesi", "Belirtilmedi")
        reason = item.get("sebep", "-")
        link = item.get("link", "")

        if link and link.startswith("http"):
            link_md = f"[Ürüne Git]({link})"
        else:
            link_md = "Link Yok"

        rows.append(f"| **{name}** | {qty} | `{price}` | {reason} | {link_md} |")

    return header + table_head + table_divider + "\n".join(rows)


def search_material_in_report(user_id: str, query: str) -> dict:
    data = load_excel_to_memory()
    if not data:
        error_msg = "Excel veri tabanı yüklenemedi."
        log_api_call(user_id, "search_material", query, error_msg)
        return {
            "status": "error",
            "message": error_msg,
            "formatted_message": f"⚠️ {error_msg}",
        }

    try:
        q_norm = normalize_text(query)
        stop_words = {
            "fiyati", "fiyat", "ne", "kadar", "var", "mi", "kac", "tl", "stok",
            "what", "is", "the", "how", "much", "nedir", "kimdir", "nasil", "listesi"
        }
        words = [w for w in q_norm.split() if w not in stop_words and len(w) > 2]

        if not words:
            not_found_msg = "Geçerli arama terimi bulunamadı."
            return {
                "status": "not_found",
                "message": not_found_msg,
                "formatted_message": f"⚠️ {not_found_msg}",
            }

        # 1. Aşama: Tam Eşleşme (AND)
        match = [
            item
            for item in data
            if all(
                w in normalize_text(f"{item['urun_adi']} {item['sebep']}")
                for w in words
            )
        ]

        # 2. Aşama: Kısmi Eşleşme (OR)
        if not match:
            match = [
                item
                for item in data
                if any(
                    w in normalize_text(f"{item['urun_adi']} {item['sebep']}")
                    for w in words
                )
            ]

        if not match:
            result_msg = f"'{query}' malzemesi mukayese raporunda bulunamadı."
            log_api_call(user_id, "search_material", query, result_msg)
            return {
                "status": "not_found",
                "message": result_msg,
                "formatted_message": f"⚠️ {result_msg}",
            }

        unique_results = list(
            {(v["urun_adi"], v["link"]): v for v in match}.values()
        )[:5]

        # Sonuçları Markdown formatına çevir
        formatted_table = format_material_response(unique_results, query)

        log_api_call(
            user_id,
            "search_material",
            query,
            json.dumps(unique_results, ensure_ascii=False),
        )
        return {
            "status": "success",
            "data": unique_results,
            "formatted_message": formatted_table,
        }

    except (ValueError, KeyError, TypeError) as e:
        error_msg = f"Arama hatası: {e!s}"
        log_api_call(user_id, "search_material", query, error_msg)
        return {
            "status": "error",
            "message": error_msg,
            "formatted_message": f"⚠️ {error_msg}",
        }