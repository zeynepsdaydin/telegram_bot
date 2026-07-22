import os
from openpyxl import load_workbook
from database import log_api_call

EXCEL_FILE = "MUKAYESE RAPORU GENEL.xlsx"

cached_data = None


def tr_lower(text):
    if not text:
        return ""
    trans = str.maketrans(
        {"İ": "i", "I": "ı", "Ğ": "ğ", "Ü": "ü", "Ş": "ş", "Ö": "ö", "Ç": "ç"}
    )
    return str(text).translate(trans).lower()


def normalize_text(text):
    text = tr_lower(text)
    trans = str.maketrans(
        {"ı": "i", "ğ": "g", "ü": "u", "ş": "s", "ö": "o", "ç": "c"}
    )
    return text.translate(trans)


def load_excel_to_memory():
    global cached_data
    if cached_data is not None:
        return cached_data

    if not os.path.exists(EXCEL_FILE):
        print(f"UYARI: {EXCEL_FILE} bulunamadı! Arama servisi çalışmayacak.")
        return None

    try:
        print("Excel verileri RAM'e yükleniyor...")
        wb = load_workbook(EXCEL_FILE, read_only=False, data_only=True)
        sheet = wb[wb.sheetnames[0]]

        all_rows = list(sheet.iter_rows(values_only=True))

        prod_idx, unit_idx, qty_idx, link_idx, reason_idx = 2, 3, 4, 5, 6
        header_row_idx = 0
        row_3 = []

        for r_idx, row in enumerate(all_rows[:10]):
            row_str = [str(cell).upper().replace("İ", "I") for cell in row if cell]
            row_concat = " ".join(row_str)
            if "URUN ADI" in row_concat or "SIRA NO" in row_concat:
                header_row_idx = r_idx
                if r_idx > 0:
                    row_3 = all_rows[r_idx - 1]
                for c_idx, cell in enumerate(row):
                    if cell:
                        c_upper = str(cell).upper().replace("İ", "I").strip()
                        if "URUN ADI" in c_upper:
                            prod_idx = c_idx
                        elif "BIRIM" in c_upper and "BIRIM FIYAT" not in c_upper:
                            unit_idx = c_idx
                        elif "MIKTAR" in c_upper:
                            qty_idx = c_idx
                        elif "LINK" in c_upper:
                            link_idx = c_idx
                        elif "SEBEP" in c_upper:
                            reason_idx = c_idx
                break

        price_columns = {}
        if header_row_idx < len(all_rows):
            header_row = all_rows[header_row_idx]
            last_found_company = "Belirtilmeyen Firma"

            for idx, cell in enumerate(header_row):
                if idx < len(row_3) and row_3[idx] is not None and str(row_3[idx]).strip():
                    last_found_company = str(row_3[idx]).strip()

                if cell:
                    c_upper = str(cell).upper().replace("İ", "I").strip()
                    if "BIRIM FIYAT" in c_upper or "FIYAT" in c_upper:
                        price_columns[idx] = last_found_company

        temp_data = []
        for row in all_rows[header_row_idx + 1 :]:
            if not row or len(row) <= prod_idx or row[prod_idx] is None:
                continue

            urun_adi_val = str(row[prod_idx]).strip()
            if not urun_adi_val or "URUN ADI" in urun_adi_val.upper() or urun_adi_val.isdigit():
                continue

            fiyat_teklifleri = []
            for col_idx, f_name in price_columns.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    val = str(row[col_idx]).strip()
                    if val and val.lower() != "none" and val != "0" and val != "0.0" and val.lower() != "stok yok":
                        if not val.endswith("TL") and "₺" not in val:
                            val = f"{val} TL"
                        fiyat_teklifleri.append(f"{f_name}: {val}")

            fiyat_ozeti = ", ".join(fiyat_teklifleri) if fiyat_teklifleri else "Fiyat belirtilmedi / Stok Yok"

            temp_data.append({
                "urun_adi": urun_adi_val,
                "birim": str(row[unit_idx]).strip() if len(row) > unit_idx and row[unit_idx] is not None else "Belirtilmedi",
                "miktar": str(row[qty_idx]).strip() if len(row) > qty_idx and row[qty_idx] is not None else "Belirtilmedi",
                "link": str(row[link_idx]).strip() if len(row) > link_idx and row[link_idx] is not None else "Belirtilmedi",
                "sebep": str(row[reason_idx]).strip() if len(row) > reason_idx and row[reason_idx] is not None else "Belirtilmedi",
                "fiyat_listesi": fiyat_ozeti,
            })

        cached_data = temp_data
        print(f"Başarıyla {len(cached_data)} satır malzeme RAM'e yüklendi!")
        if cached_data:
            print(f"[EXCEL DENEME OKUMA] İlk 3 Ürün Adı: {[item['urun_adi'] for item in cached_data[:3]]}")
        return cached_data
    except Exception as e:
        print(f"Excel RAM'e yüklenirken hata oluştu: {e}")
        return None


load_excel_to_memory()


def search_material_in_report(user_id, query):
    data = load_excel_to_memory()
    if not data:
        error_msg = "Hata: Excel veri tabanı yüklü değil."
        log_api_call(user_id, "search_material", query, error_msg)
        return {"error": error_msg}

    try:
        q_norm = normalize_text(query)
        words = [w for w in q_norm.split() if w not in ["fiyati", "fiyat", "ne", "kadar", "var", "mi", "kac"] and len(w) > 1]

        print("\n--- [ARAMA TETİKLENDİ] ---")
        print(f"Gelen Sorgu: '{query}'")
        print(f"Arama Kelimeleri: {words}")

        match = []
        for item in data:
            p_norm = normalize_text(item["urun_adi"])
            p_full = normalize_text(f"{item['urun_adi']} {item['sebep']} {item['link']}")

            if any(w in p_norm for w in words) or any(w in p_full for w in words):
                match.append(item)

        print(f"Bulunan Eşleşme Sayısı: {len(match)}")
        if match:
            print(f"İlk Eşleşen Ürün: {match[0]['urun_adi']}")
        print("---------------------------\n")

        if not match:
            result_msg = f"Aradığınız '{query}' malzemesi mukayese raporunda bulunamadı."
            log_api_call(user_id, "search_material", query, result_msg)
            return {"status": "not_found", "message": result_msg}

        unique_results = list({v['urun_adi']: v for v in match}.values())
        results = unique_results[:5]

        log_api_call(user_id, "search_material", query, str(results))
        return {"status": "success", "data": results}
    except Exception as e:
        error_msg = f"Sorgu sırasında hata oluştu: {str(e)}"
        log_api_call(user_id, "search_material", query, error_msg)
        return {"error": error_msg}